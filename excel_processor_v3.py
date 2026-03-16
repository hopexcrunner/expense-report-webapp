from openpyxl import load_workbook
from datetime import datetime
import io, os, requests

class ExcelProcessorV3:
    def __init__(self):
        self.template_path = "Avant_2026_Expense_Report_Form.xlsx"
        self.CATEGORY_ROWS = {
            'Travel': 42,
            'Advance': 37,
            'Ministry Entertainment': 61,
            'Other': 111,
            'Honorariums': 165
        }
    
    def get_exchange_rate(self, from_currency, to_currency, date_str):
        """Get exchange rate using European Central Bank API (via frankfurter.app)"""
        if from_currency == to_currency:
            return 1.0
        
        # Parse date
        date_obj = None
        for fmt in ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d']:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                break
            except:
                continue
        
        if not date_obj:
            date_obj = datetime.now()
        
        # Try European Central Bank API via frankfurter.app (free ECB data)
        try:
            api_date = date_obj.strftime('%Y-%m-%d')
            url = f"https://api.frankfurter.app/{api_date}"
            params = {'from': from_currency, 'to': to_currency}
            response = requests.get(url, params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                if 'rates' in data and to_currency in data['rates']:
                    rate = float(data['rates'][to_currency])
                    print(f"✅ ECB rate for {date_str}: {from_currency}→{to_currency} = {rate}")
                    return rate
        except Exception as e:
            print(f"⚠️ ECB API unavailable: {e}")
        
        # Fallback: Use approximate monthly rates (updated January 2026)
        # Based on average ECB rates - update quarterly
        fallback_rates = {
            ('EUR', 'USD'): 1.12,  # Updated from actual Jan 2026 rates
            ('USD', 'EUR'): 0.89,
            ('GBP', 'USD'): 1.27,
            ('GBP', 'EUR'): 1.17,
            ('USD', 'GBP'): 0.79,
            ('EUR', 'GBP'): 0.85
        }
        
        fallback = fallback_rates.get((from_currency, to_currency), 1.0)
        print(f"⚠️ Using fallback rate for {date_str}: {from_currency}→{to_currency} = {fallback}")
        return fallback
    
    def create_expense_report(self, employee_info, receipts):
        if not os.path.exists(self.template_path):
            return self._create_simple_report(employee_info, receipts)
        
        wb = load_workbook(self.template_path)
        sheet = wb['Expense Report']
        
        # Employee info - write to FIRST cell of merged range
        print("\n📝 Writing employee info...")
        self._write_to_merged_cell(sheet, 6, 3, employee_info.get('name', ''), 'Name')
        self._write_to_merged_cell(sheet, 6, 7, employee_info.get('date_submitted', ''), 'Date Submitted')
        
        if employee_info.get('account_project'):
            self._write_to_merged_cell(sheet, 9, 3, employee_info.get('account_project'), 'Account/Project')
        if employee_info.get('field'):
            self._write_to_merged_cell(sheet, 9, 7, employee_info.get('field'), 'Field')
        
        # Signature in ROW 24
        self._write_to_merged_cell(sheet, 24, 3, employee_info.get('name', ''), 'Worker Signature')
        self._write_to_merged_cell(sheet, 24, 7, employee_info.get('signature_date', ''), 'Signature Date')
        
        # Process receipts
        print("\n📝 Processing receipts...")
        category_counters = {cat: 0 for cat in self.CATEGORY_ROWS.keys()}
        
        for idx, receipt in enumerate(receipts, 1):
            category = receipt.get('category', 'Travel')
            start_row = self.CATEGORY_ROWS.get(category, 42)
            row = start_row + category_counters[category]
            category_counters[category] += 1
            
            if row > 200:
                continue
            
            print(f"\nReceipt {idx} ({category}):")
            
            # Get exchange rate using RECEIPT DATE
            currency = receipt.get('currency', 'EUR')
            receipt_date = receipt['data'].get('date', '')
            exchange_rate = self.get_exchange_rate(currency, 'USD', receipt_date)
            
            # Date in MM/DD/YYYY
            self._set_cell_value_safe(sheet, row, 2, receipt['data'].get('date', ''))
            
            if category == 'Travel':
                self._set_cell_value_safe(sheet, row, 3, receipt.get('from_location', ''))
                self._set_cell_value_safe(sheet, row, 4, receipt.get('to_location', ''))
            elif category == 'Ministry Entertainment':
                self._set_cell_value_safe(sheet, row, 3, receipt.get('who', ''))
                self._set_cell_value_safe(sheet, row, 4, receipt.get('where', ''))
            else:
                self._set_cell_value_safe(sheet, row, 3, receipt['data'].get('merchant', ''))
            
            self._set_cell_value_safe(sheet, row, 5, receipt.get('ministry_purpose', ''))
            self._set_cell_value_safe(sheet, row, 6, receipt['data'].get('total', 0.0))
            self._set_cell_value_safe(sheet, row, 7, exchange_rate)
            
            print(f"  ✅ Written to row {row}")
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _write_to_merged_cell(self, sheet, row, col, value, label):
        """Write to first cell of a merged range"""
        try:
            cell = sheet.cell(row=row, column=col)
            coord = cell.coordinate
            
            # Check if this cell is part of a merged range
            for merged_range in sheet.merged_cells.ranges:
                if coord in merged_range:
                    # Get the top-left cell of the merged range
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    top_left_cell = sheet.cell(min_row, min_col)
                    top_left_cell.value = value
                    print(f"  ✅ {label}: '{value}' → {top_left_cell.coordinate}")
                    return
            
            # Not merged, write normally
            cell.value = value
            print(f"  ✅ {label}: '{value}' → {coord}")
            
        except Exception as e:
            print(f"  ❌ Error writing {label}: {e}")
    
    def _set_cell_value_safe(self, sheet, row, col, value):
        """Set cell value, skipping merged cells in receipt rows"""
        try:
            cell = sheet.cell(row=row, column=col)
            coord = cell.coordinate
            
            # Check if this is a merged cell
            for merged_range in sheet.merged_cells.ranges:
                if coord in merged_range:
                    # Skip merged cells in receipt rows
                    return
            
            # Safe to write
            if isinstance(value, (int, float)):
                cell.value = value
            elif isinstance(value, str):
                cell.value = value
            else:
                cell.value = str(value) if value else ''
                
        except Exception:
            pass
    
    def _create_simple_report(self, employee_info, receipts):
        from openpyxl import Workbook
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'Avant Expense Report'
        sheet['A3'] = employee_info.get('name', '')
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
