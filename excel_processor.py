from openpyxl import load_workbook
from datetime import datetime
import io
import os
import requests

class ExcelProcessor:
    """Process Excel expense report templates"""
    
    def __init__(self):
        self.template_path = "Avant_2026_Expense_Report_Form.xlsx"
    
    def get_exchange_rate(self, from_currency, to_currency, date_str=None):
        """
        Get exchange rate for a specific date
        
        Args:
            from_currency: Source currency (e.g., 'EUR')
            to_currency: Target currency (e.g., 'USD')
            date_str: Date string in format DD/MM/YYYY or None for today
            
        Returns:
            float: Exchange rate or 1.0 if same currency
        """
        if from_currency == to_currency:
            return 1.0
        
        try:
            # Parse date if provided
            if date_str:
                try:
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                except:
                    date_obj = datetime.now()
            else:
                date_obj = datetime.now()
            
            # Use exchangerate.host API (free, no key needed)
            date_formatted = date_obj.strftime('%Y-%m-%d')
            url = f"https://api.exchangerate.host/{date_formatted}"
            params = {
                'base': from_currency,
                'symbols': to_currency
            }
            
            response = requests.get(url, params=params, timeout=5)
            
            if response.status_code == 200:
                data = response.json()
                if 'rates' in data and to_currency in data['rates']:
                    return float(data['rates'][to_currency])
        
        except Exception as e:
            print(f"Exchange rate fetch failed: {e}")
        
        # Fallback rates (approximate)
        fallback_rates = {
            ('EUR', 'USD'): 1.08,
            ('USD', 'EUR'): 0.93,
            ('GBP', 'USD'): 1.27,
            ('USD', 'GBP'): 0.79,
        }
        
        return fallback_rates.get((from_currency, to_currency), 1.0)
    
    def create_expense_report(self, receipt_data):
        """
        Create expense report from receipt data
        
        Args:
            receipt_data: Dictionary with merchant, date, total, items, etc.
            
        Returns:
            BytesIO buffer with Excel file
        """
        # Load template
        if os.path.exists(self.template_path):
            wb = load_workbook(self.template_path)
        else:
            # Create simple workbook if template doesn't exist
            return self._create_simple_report(receipt_data)
        
        sheet = wb['Expense Report']
        
        # Fill header information
        # Date submitted (G7)
        current_date = datetime.now().strftime('%m/%d/%Y')
        self._set_cell_value(sheet, 7, 7, current_date)  # Row 7, Column G
        
        # Get exchange rate
        currency = receipt_data.get('currency', 'EUR')
        receipt_date = receipt_data.get('date', '')
        exchange_rate = self.get_exchange_rate(currency, 'USD', receipt_date)
        
        # Add expense line items starting at row 42
        start_row = 42
        items = receipt_data.get('items', [])
        
        if items:
            # Add each line item
            for i, item in enumerate(items):
                row = start_row + i
                if row < 129:  # Stay within expense area
                    # Date (Column B)
                    self._set_cell_value(sheet, row, 2, receipt_data.get('date', ''))
                    
                    # Merchant (Column C - From)
                    self._set_cell_value(sheet, row, 3, receipt_data.get('merchant', ''))
                    
                    # Address (Column D - To)
                    self._set_cell_value(sheet, row, 4, receipt_data.get('address', ''))
                    
                    # Description (Column E - Purpose)
                    self._set_cell_value(sheet, row, 5, item.get('description', ''))
                    
                    # Amount (Column F - Local Currency)
                    self._set_cell_value(sheet, row, 6, item.get('amount', 0.0))
                    
                    # Exchange Rate (Column G)
                    self._set_cell_value(sheet, row, 7, exchange_rate)
                    
                    # Account Code (Column M)
                    self._set_cell_value(sheet, row, 13, "8464 - Meals & Entertainment")
        else:
            # Single entry for total amount
            row = start_row
            self._set_cell_value(sheet, row, 2, receipt_data.get('date', ''))
            self._set_cell_value(sheet, row, 3, receipt_data.get('merchant', ''))
            self._set_cell_value(sheet, row, 4, receipt_data.get('address', ''))
            self._set_cell_value(sheet, row, 5, "Expense")
            self._set_cell_value(sheet, row, 6, receipt_data.get('total', 0.0))
            self._set_cell_value(sheet, row, 7, exchange_rate)
            self._set_cell_value(sheet, row, 13, "8464 - Meals & Entertainment")
        
        # Save to BytesIO buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _set_cell_value(self, sheet, row, col, value):
        """Set cell value handling different types"""
        cell = sheet.cell(row=row, column=col)
        
        if isinstance(value, (int, float)):
            cell.value = value
        elif isinstance(value, str):
            cell.value = value
        else:
            cell.value = str(value)
    
    def _create_simple_report(self, receipt_data):
        """Create a simple Excel report if template is not available"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Expense Report"
        
        # Header
        sheet['A1'] = 'Avant Expense Report'
        sheet['A1'].font = Font(size=16, bold=True)
        
        # Date
        sheet['A3'] = 'Date Submitted:'
        sheet['B3'] = datetime.now().strftime('%m/%d/%Y')
        
        # Receipt info
        sheet['A5'] = 'Merchant:'
        sheet['B5'] = receipt_data.get('merchant', '')
        
        sheet['A6'] = 'Date:'
        sheet['B6'] = receipt_data.get('date', '')
        
        sheet['A7'] = 'Total Amount:'
        sheet['B7'] = receipt_data.get('total', 0.0)
        sheet['B7'].number_format = '€#,##0.00'
        
        # Line items header
        sheet['A9'] = 'Description'
        sheet['B9'] = 'Quantity'
        sheet['C9'] = 'Unit Price'
        sheet['D9'] = 'Amount'
        
        # Style header
        for cell in ['A9', 'B9', 'C9', 'D9']:
            sheet[cell].font = Font(bold=True)
            sheet[cell].fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Add items
        items = receipt_data.get('items', [])
        row = 10
        
        if items:
            for item in items:
                sheet[f'A{row}'] = item.get('description', '')
                sheet[f'B{row}'] = item.get('quantity', 1)
                sheet[f'C{row}'] = item.get('unit_price', 0.0)
                sheet[f'D{row}'] = item.get('amount', 0.0)
                
                # Format currency
                sheet[f'C{row}'].number_format = '€#,##0.00'
                sheet[f'D{row}'].number_format = '€#,##0.00'
                
                row += 1
        else:
            sheet[f'A{row}'] = 'Single Expense'
            sheet[f'B{row}'] = 1
            sheet[f'C{row}'] = receipt_data.get('total', 0.0)
            sheet[f'D{row}'] = receipt_data.get('total', 0.0)
            
            sheet[f'C{row}'].number_format = '€#,##0.00'
            sheet[f'D{row}'].number_format = '€#,##0.00'
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
