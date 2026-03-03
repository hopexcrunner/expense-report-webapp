from openpyxl import load_workbook
from datetime import datetime
import io
import os
import requests

class ExcelProcessorV2:
    """Excel processor with correct row mapping"""
    
    def __init__(self):
        self.template_path = "Avant_2026_Expense_Report_Form.xlsx"
        self.CATEGORY_ROWS = {
            'Travel': 42,
            'Advance': 37,
            'Ministry Entertainment': 64,
            'Other': 75,
            'Honorariums': 165
        }
    
    def get_exchange_rate(self, from_currency, to_currency, date_str=None):
        if from_currency == to_currency:
            return 1.0
        try:
            if date_str:
                try:
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                except:
                    date_obj = datetime.now()
            else:
                date_obj = datetime.now()
            
            date_formatted = date_obj.strftime('%Y-%m-%d')
            url = f"https://api.exchangerate.host/{date_formatted}"
            params = {'base': from_currency, 'symbols': to_currency}
            response = requests.get(url, params=params, timeout=5)
            
            if response.status_code == 200:
                data = response.json()
                if 'rates' in data and to_currency in data['rates']:
                    return float(data['rates'][to_currency])
        except Exception:
            pass
        
        fallback_rates = {('EUR', 'USD'): 1.08, ('USD', 'EUR'): 0.93, ('GBP', 'USD'): 1.27}
        return fallback_rates.get((from_currency, to_currency), 1.0)
    
    def create_expense_report(self, employee_info, receipts):
        if os.path.exists(self.template_path):
            wb = load_workbook(self.template_path)
        else:
            return self._create_simple_report(employee_info, receipts)
        
        sheet = wb['Expense Report']
        
        # Employee info - Row 6 NOT Row 7
        self._set_cell_value(sheet, 6, 3, employee_info.get('name', ''))
        self._set_cell_value(sheet, 6, 7, employee_info.get('date_submitted', ''))
        
        if employee_info.get('account_project'):
            self._set_cell_value(sheet, 9, 3, employee_info.get('account_project'))
        if employee_info.get('field'):
            self._set_cell_value(sheet, 9, 7, employee_info.get('field'))
        
        self._set_cell_value(sheet, 11, 3, employee_info.get('name', ''))
        self._set_cell_value(sheet, 11, 7, employee_info.get('signature_date', ''))
        
        # Process receipts
        category_counters = {cat: 0 for cat in self.CATEGORY_ROWS.keys()}
        
        for receipt in receipts:
            category = receipt.get('category', 'Travel')
            start_row = self.CATEGORY_ROWS.get(category, 42)
            row = start_row + category_counters[category]
            category_counters[category] += 1
            
            if row > 200:
                continue
            
            currency = receipt['data'].get('currency', 'EUR')
            receipt_date = receipt['data'].get('date', '')
            exchange_rate = self.get_exchange_rate(currency, 'USD', receipt_date)
            
            self._set_cell_value(sheet, row, 2, receipt['data'].get('date', ''))
            
            if category == 'Travel':
                self._set_cell_value(sheet, row, 3, receipt.get('from_location', ''))
                self._set_cell_value(sheet, row, 4, receipt.get('to_location', ''))
            else:
                self._set_cell_value(sheet, row, 3, receipt['data'].get('merchant', ''))
                self._set_cell_value(sheet, row, 4, receipt['data'].get('address', ''))
            
            self._set_cell_value(sheet, row, 5, receipt.get('ministry_purpose', ''))
            self._set_cell_value(sheet, row, 6, receipt['data'].get('total', 0.0))
            self._set_cell_value(sheet, row, 7, exchange_rate)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _set_cell_value(self, sheet, row, col, value):
        cell = sheet.cell(row=row, column=col)
        if isinstance(value, (int, float)):
            cell.value = value
        elif isinstance(value, str):
            cell.value = value
        else:
            cell.value = str(value) if value else ''
    
    def _create_simple_report(self, employee_info, receipts):
        from openpyxl import Workbook
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'Avant Expense Report (Simple)'
        sheet['A3'] = employee_info.get('name', '')
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
